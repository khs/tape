"""
Build the CBSA crosswalk files under pipelines/_crosswalks/ from
Census's published CBSA delineation file.

Writes two CSVs in one run:

  * cbsa_metro.csv       — one row per MSA, columns:
        cbsa_code, short_name, name, primary_state_fips,
        primary_state_abbr, states

  * county_to_cbsa.csv   — one row per (county, MSA) pair, columns:
        county_fips, cbsa_code, county_name, state_abbr
    Used by pipelines/usaspending_metro.py to aggregate county-level
    federal-spending rows up to the MSA.

The repo ships a curated starter cbsa_metro.csv (~100 of the largest
MSAs) so the Metro pipelines have something to chew on out of the box.
Run this script to overwrite both with the full ~387-MSA / ~1100-county
OMB Bulletin list.

Source: Census Bureau "Delineation Files for Metropolitan and
Micropolitan Statistical Areas." We pull the latest March-vintage Excel
release. The exact URL changes per OMB bulletin revision; bump
CBSA_LIST_URL when OMB issues a new bulletin.

`primary_state` = the state listed first in the CBSA name (the "lead"
state by convention). `states` = pipe-separated list of every state the
CBSA touches.

Run with: ``python pipelines/build_cbsa_metro_crosswalk.py``.

Requires openpyxl (``pip install openpyxl``).
"""
from __future__ import annotations

import csv
import subprocess
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
CROSSWALK_DIR = REPO_ROOT / "pipelines" / "_crosswalks"
METRO_OUT_PATH = CROSSWALK_DIR / "cbsa_metro.csv"
COUNTY_OUT_PATH = CROSSWALK_DIR / "county_to_cbsa.csv"

# Census publishes the CBSA delineation list as both XLS and a flat-text
# substitute. We pull the CSV-friendly "list1" sheet. As of OMB Bulletin
# 23-01 (July 2023):
CBSA_LIST_URL = (
    "https://www2.census.gov/programs-surveys/metro-micro/geographies/"
    "reference-files/2023/delineation-files/list1_2023.xlsx"
)

# State FIPS → abbreviation (same table as the other pipelines).
STATE_ABBR: dict[str, str] = {
    "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA", "08": "CO",
    "09": "CT", "10": "DE", "11": "DC", "12": "FL", "13": "GA", "15": "HI",
    "16": "ID", "17": "IL", "18": "IN", "19": "IA", "20": "KS", "21": "KY",
    "22": "LA", "23": "ME", "24": "MD", "25": "MA", "26": "MI", "27": "MN",
    "28": "MS", "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND", "39": "OH",
    "40": "OK", "41": "OR", "42": "PA", "44": "RI", "45": "SC", "46": "SD",
    "47": "TN", "48": "TX", "49": "UT", "50": "VT", "51": "VA", "53": "WA",
    "54": "WV", "55": "WI", "56": "WY", "72": "PR",
}
ABBR_TO_FIPS = {abbr: fips for fips, abbr in STATE_ABBR.items()}


def fetch_xlsx(url: str) -> bytes:
    """Download the CBSA delineation XLSX. Raises on non-2xx."""
    try:
        out = subprocess.run(
            ["curl", "-sSL", "--max-time", "120", "-o", "-", url],
            capture_output=True, check=True,
        )
    except subprocess.CalledProcessError as e:
        print(f"  curl failed: {e}", file=sys.stderr)
        raise
    return out.stdout


def parse_xlsx(blob: bytes) -> list[dict[str, str]]:
    """
    Parse Census's list1 sheet. Uses openpyxl if available; otherwise
    raises with install instructions.

    Returns one row per (CBSA, county) pair — caller aggregates to one
    row per CBSA.
    """
    try:
        from openpyxl import load_workbook  # noqa: WPS433 (local import)
    except ImportError:
        raise SystemExit(
            "openpyxl not installed. Run:\n"
            "    pip install openpyxl\n"
            "then re-run this script."
        )
    import io
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active

    # Find the header row — Census prepends 2 title rows.
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows[:10]):
        if row and any(cell == "CBSA Code" for cell in row if cell is not None):
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit("Could not find header row in CBSA list1 sheet.")
    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]

    out: list[dict[str, str]] = []
    for r in rows[header_idx + 1:]:
        if not r or all(c is None for c in r):
            continue
        rec = {headers[i]: ("" if r[i] is None else str(r[i]).strip())
               for i in range(min(len(headers), len(r)))}
        if not rec.get("CBSA Code"):
            continue
        out.append(rec)
    return out


def aggregate(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """
    Collapse the per-county rows down to one row per CBSA. Restrict to
    Metropolitan Statistical Areas (not Micropolitan).
    """
    by_cbsa: dict[str, dict[str, object]] = {}
    for rec in rows:
        kind = rec.get("Metropolitan/Micropolitan Statistical Area", "")
        if "Metropolitan" not in kind:
            continue
        code = rec["CBSA Code"]
        if code not in by_cbsa:
            full_name = rec.get("CBSA Title", "")
            # Heuristic: "Atlanta-Sandy Springs-Alpharetta, GA" → short =
            # everything before the comma; primary state = the first
            # state abbr after the comma.
            short = full_name.split(",")[0].strip()
            state_tail = full_name.split(",", 1)[1].strip() if "," in full_name else ""
            primary_abbr = state_tail.split("-")[0].strip() if state_tail else ""
            by_cbsa[code] = {
                "cbsa_code": code,
                "name": full_name,
                "short_name": short,
                "primary_state_abbr": primary_abbr,
                "primary_state_fips": ABBR_TO_FIPS.get(primary_abbr, ""),
                "_state_set": set(),
            }
        state_abbr = rec.get("State Name", "")
        # The XLSX has both "State Name" (Alabama) and a FIPS column. We
        # want the abbreviation; derive from FIPS to avoid name-matching.
        fips = rec.get("FIPS State Code", "")
        if fips and fips.zfill(2) in STATE_ABBR:
            cast = by_cbsa[code]["_state_set"]
            assert isinstance(cast, set)
            cast.add(STATE_ABBR[fips.zfill(2)])

    final: list[dict[str, str]] = []
    for code, agg in sorted(by_cbsa.items()):
        state_set = agg.pop("_state_set")
        assert isinstance(state_set, set)
        # Order states by the order they appear in the CBSA title (so
        # "NY-NJ-PA" stays that way), then anything else alphabetically.
        title_states: list[str] = []
        tail = agg["name"].split(",", 1)[1].strip() if "," in str(agg["name"]) else ""
        for tok in tail.split("-"):
            tok = tok.strip()
            if tok in state_set and tok not in title_states:
                title_states.append(tok)
        for s in sorted(state_set):
            if s not in title_states:
                title_states.append(s)
        agg["states"] = "|".join(title_states)
        final.append({k: str(v) for k, v in agg.items()})
    return final


def write_metro_csv(rows: list[dict[str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["cbsa_code", "short_name", "name", "primary_state_fips",
            "primary_state_abbr", "states"]
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})


def write_county_csv(raw_rows: list[dict[str, str]], path: Path) -> int:
    """
    One row per (county, MSA) pair. Restricted to Metropolitan
    Statistical Areas; Micropolitan counties don't roll up to MSAs and
    would just add noise for the spending pipeline.

    Returns the number of rows written.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["county_fips", "cbsa_code", "county_name", "state_abbr"]
    written = 0
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for rec in raw_rows:
            kind = rec.get("Metropolitan/Micropolitan Statistical Area", "")
            if "Metropolitan" not in kind:
                continue
            state_fips = (rec.get("FIPS State Code") or "").zfill(2)
            county_fips_3 = (rec.get("FIPS County Code") or "").zfill(3)
            cbsa = (rec.get("CBSA Code") or "").strip()
            if not state_fips or not county_fips_3 or not cbsa:
                continue
            # 5-digit state+county FIPS — the form USAspending's
            # county shape_code uses.
            fips5 = state_fips + county_fips_3
            w.writerow({
                "county_fips": fips5,
                "cbsa_code": cbsa,
                "county_name": rec.get("County/County Equivalent", "").strip(),
                "state_abbr": STATE_ABBR.get(state_fips, ""),
            })
            written += 1
    return written


def main() -> int:
    print(f"Fetching {CBSA_LIST_URL}...", flush=True)
    blob = fetch_xlsx(CBSA_LIST_URL)
    print(f"  got {len(blob):,} bytes", flush=True)
    raw = parse_xlsx(blob)
    print(f"  parsed {len(raw):,} county-level rows", flush=True)

    metros = aggregate(raw)
    print(f"  collapsed to {len(metros):,} Metropolitan Statistical Areas",
          flush=True)
    write_metro_csv(metros, METRO_OUT_PATH)
    print(f"  wrote {METRO_OUT_PATH.relative_to(REPO_ROOT)}", flush=True)

    county_rows = write_county_csv(raw, COUNTY_OUT_PATH)
    # ASCII '->' rather than '→' so Windows' default cp1252
    # stdout codec doesn't barf at print time.
    print(f"  wrote {COUNTY_OUT_PATH.relative_to(REPO_ROOT)} "
          f"({county_rows:,} county -> MSA mappings)", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
