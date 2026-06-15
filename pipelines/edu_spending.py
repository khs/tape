"""
K-12 current spending per pupil, by state + US — U.S. Census Bureau
Annual Survey of School System Finances (the NCES/Census F-33 program),
summary Table 8 ("Per Pupil Amounts for Current Spending of Public
Elementary-Secondary School Systems by State").

Per-pupil current spending = total current expenditure on elementary-
secondary education divided by enrollment; the standard cross-state
school-funding comparison. We source it DIRECTLY from the Census summary
tables so the entire series uses one consistent methodology.

History: this was previously sourced via the Urban Institute Education
Data Portal, which (a) caps at school year 2019-20 and (b) computes
per-pupil with a ~5%-different enrollment denominator than the Census
per-pupil table — so splicing newer Census years onto the Urban series
would have put a false ~5% kink at the 2020 seam (Urban FY2020 = $14,147
vs Census FY2020 = $13,501). Census Table 8 is the authoritative figure;
its per-state values match the Census per-pupil press releases to the
dollar (e.g. FY2022 NY $29,873, UT $9,552, MS $10,984).

Source format: Census publishes only Excel summary tables (no API), at
  www2.census.gov/programs-surveys/school-finances/tables/<year>/
    secondary-education-finance/elsec<YY>_sumtables<ext>
<ext> is .xls for FY<=2021 and .xlsx for FY>=2022; both carry the same
Table 8 layout (sheet "8", column C = "Total" per-pupil current spending,
rows = US + each state). We parse both formats and validate the US value
sits in a sane band so a silent upstream structure change fails loudly
rather than shipping wrong numbers.

Run: `python pipelines/edu_spending.py` (no key; public Census files).
Requires openpyxl (.xlsx) + xlrd (.xls).
"""
from __future__ import annotations

import io
import sys
import urllib.request
from datetime import datetime

import _env  # noqa: F401
import openpyxl
import xlrd

from common import write_timeseries
from pathlib import Path

PIPELINE = "edu_spending"
BASE = (
    "https://www2.census.gov/programs-surveys/school-finances/tables/"
    "{year}/secondary-education-finance/elsec{yy}_sumtables{ext}"
)
START_YEAR = 2010
# Table 8: column C (0-indexed 2) is "Total" per-pupil current spending,
# in every observed year and both formats. Values sit in this band; we
# validate against it so a layout change can't silently ship wrong data.
PERPUPIL_COL = 2
SANE_MIN, SANE_MAX = 5_000, 60_000
MIN_STATES = 40  # a real Table 8 has US + 50 states + DC

# fips -> (abbr, name). FIPS is unused now (Census parses by name) but the
# abbr/name mapping drives the series IDs + YAML.
STATES: dict[int, tuple[str, str]] = {
    1: ("AL", "Alabama"), 2: ("AK", "Alaska"), 4: ("AZ", "Arizona"),
    5: ("AR", "Arkansas"), 6: ("CA", "California"), 8: ("CO", "Colorado"),
    9: ("CT", "Connecticut"), 10: ("DE", "Delaware"),
    11: ("DC", "District of Columbia"), 12: ("FL", "Florida"),
    13: ("GA", "Georgia"), 15: ("HI", "Hawaii"), 16: ("ID", "Idaho"),
    17: ("IL", "Illinois"), 18: ("IN", "Indiana"), 19: ("IA", "Iowa"),
    20: ("KS", "Kansas"), 21: ("KY", "Kentucky"), 22: ("LA", "Louisiana"),
    23: ("ME", "Maine"), 24: ("MD", "Maryland"), 25: ("MA", "Massachusetts"),
    26: ("MI", "Michigan"), 27: ("MN", "Minnesota"), 28: ("MS", "Mississippi"),
    29: ("MO", "Missouri"), 30: ("MT", "Montana"), 31: ("NE", "Nebraska"),
    32: ("NV", "Nevada"), 33: ("NH", "New Hampshire"), 34: ("NJ", "New Jersey"),
    35: ("NM", "New Mexico"), 36: ("NY", "New York"), 37: ("NC", "North Carolina"),
    38: ("ND", "North Dakota"), 39: ("OH", "Ohio"), 40: ("OK", "Oklahoma"),
    41: ("OR", "Oregon"), 42: ("PA", "Pennsylvania"), 44: ("RI", "Rhode Island"),
    45: ("SC", "South Carolina"), 46: ("SD", "South Dakota"), 47: ("TN", "Tennessee"),
    48: ("TX", "Texas"), 49: ("UT", "Utah"), 50: ("VT", "Vermont"),
    51: ("VA", "Virginia"), 53: ("WA", "Washington"), 54: ("WV", "West Virginia"),
    55: ("WI", "Wisconsin"), 56: ("WY", "Wyoming"),
}
NAME_TO_ABBR = {name: abbr for abbr, name in STATES.values()}
STATE_NAME_LC = {abbr.lower(): name for abbr, name in STATES.values()}

REPO_ROOT = Path(__file__).resolve().parent.parent
YAML_DIR = REPO_ROOT / "src" / "content" / "sources" / PIPELINE
CENSUS_URL = "https://www.census.gov/programs-surveys/school-finances.html"
LICENSE = "Public domain (US Census Bureau)"


def _q(s: str) -> str:
    return "'" + s.replace("'", "''") + "'"


def _clean_geo(a: object) -> str:
    """Strip the dotted padding / footnote markers from a Table 8 area cell."""
    if not isinstance(a, str):
        return ""
    return " ".join("".join(c for c in a if c.isalpha() or c == " ").split())


def _fetch(url: str) -> bytes | None:
    try:
        with urllib.request.urlopen(url, timeout=120) as r:  # noqa: S310
            if getattr(r, "status", 200) != 200:
                return None
            return r.read()
    except Exception:  # noqa: BLE001 — 404 / network: just try the next variant
        return None


def _table8_rows(data: bytes, ext: str) -> list[tuple[object, object]]:
    """Return [(area_cell, perpupil_cell), ...] from Table 8 of a summary file."""
    out: list[tuple[object, object]] = []
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        if "8" not in wb.sheetnames:
            return []
        for row in wb["8"].iter_rows(values_only=True):
            if len(row) > PERPUPIL_COL:
                out.append((row[0], row[PERPUPIL_COL]))
    else:  # .xls
        wb = xlrd.open_workbook(file_contents=data)
        if "8" not in wb.sheet_names():
            return []
        sh = wb.sheet_by_name("8")
        for r in range(sh.nrows):
            v = sh.cell_value(r, PERPUPIL_COL) if sh.ncols > PERPUPIL_COL else None
            out.append((sh.cell_value(r, 0), v))
    return out


def fetch_year(year: int) -> dict[str, float]:
    """{geo_key: per_pupil} for one fiscal year ('us' + state abbrs), or {}."""
    yy = f"{year % 100:02d}"
    for ext in (".xlsx", ".xls"):
        data = _fetch(BASE.format(year=year, yy=yy, ext=ext))
        if not data:
            continue
        try:
            rows = _table8_rows(data, ext)
        except Exception:  # noqa: BLE001 — wrong format / corrupt: try next
            continue
        result: dict[str, float] = {}
        for area, v in rows:
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            if not (SANE_MIN <= v <= SANE_MAX):
                continue
            name = _clean_geo(area)
            if name == "United States":
                result["us"] = round(float(v), 2)
            elif name in NAME_TO_ABBR:
                result[NAME_TO_ABBR[name].lower()] = round(float(v), 2)
        # Structure sanity: a real Table 8 yields the US row + ~all states.
        if "us" in result and len(result) >= MIN_STATES:
            return result
    return {}


def write_yaml(sid: str, geo_key: str) -> None:
    if geo_key == "us":
        name = "K-12 spending per pupil (US)"
        short = "US per-pupil K-12"
        where = "the US"
        tags = ["education", "fiscal", "us"]
    else:
        where = STATE_NAME_LC[geo_key]
        name = f"{where} K-12 spending per pupil"
        short = f"{geo_key.upper()} per-pupil K-12"
        tags = ["education", "fiscal", "us-state", "us"]
    desc = (
        f"Current spending per pupil on K-12 public education in {where}, "
        f"total current expenditure divided by enrollment. US Census Bureau "
        f"Annual Survey of School System Finances (the NCES / Census F-33 "
        f"program), summary Table 8."
    )
    lines = [
        f"name: {_q(name)}",
        f"shortName: {_q(short)}",
        f"description: {_q(desc)}",
        "kind: timeseries",
        f"pipeline: {PIPELINE}",
        f"dataFile: data/{PIPELINE}/{sid}.json",
        'supportedDeltas: ["1y", "5y", "10y"]',
        'unit: "USD"',
        "emphasis: level",
        "formatting:",
        "  style: currency",
        "  currency: USD",
        "  decimals: 0",
        "provenance:",
        "  provider: U.S. Census Bureau, Annual Survey of School System Finances",
        "  series: Public Elementary-Secondary Education Finance Data, Table 8 (per-pupil current spending)",
        f"  url: {CENSUS_URL}",
        f"  license: {_q(LICENSE)}",
        "tags:",
    ]
    for t in tags:
        lines.append(f"  - {t}")
    lines.append("unitClass: currency")
    (YAML_DIR / f"{sid}.yaml").write_text("\n".join(lines) + "\n", encoding="utf-8")


def run() -> int:
    YAML_DIR.mkdir(parents=True, exist_ok=True)
    by_geo: dict[str, list[dict]] = {}
    this_year = datetime.now().year
    got_years = 0
    for year in range(START_YEAR, this_year + 1):
        res = fetch_year(year)
        if res:
            got_years += 1
        for geo_key, v in res.items():
            by_geo.setdefault(geo_key, []).append({"t": f"{year}-06-30", "v": v})

    if got_years == 0:
        print("edu_spending: no Census summary tables fetched — leaving data untouched.",
              file=sys.stderr)
        return 1

    written = 0
    for geo_key, pts in by_geo.items():
        if len(pts) < 2:
            continue
        pts.sort(key=lambda p: p["t"])
        sid = "us_perpupil" if geo_key == "us" else f"state_perpupil_{geo_key}"
        # merge=False: REPLACE, never union. The prior data was Urban-sourced
        # with a ~5%-different methodology; merging would mix the two and
        # re-create the very kink this re-source removes.
        write_timeseries(PIPELINE, sid, sid, pts, unit="USD", merge=False)
        write_yaml(sid, geo_key)
        written += 1

    print(f"edu_spending: wrote {written} per-pupil series from Census "
          f"({got_years} fiscal years, {START_YEAR}-{this_year}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
